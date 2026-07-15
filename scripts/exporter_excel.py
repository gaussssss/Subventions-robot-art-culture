"""Essai visuel du pipeline : collecte complète → fichier Excel local.

Produit sortie/veille-subventions-AAAA-MM-JJ.xlsx avec les mêmes colonnes que
le futur Google Sheet (onglets Subventions + Journal), pour vérifier de ses
yeux ce que donne la veille avant de brancher Google Sheets.

Usage :
    python scripts/exporter_excel.py                     # collecte fraîche (toutes les sources)
    python scripts/exporter_excel.py --sources calq-organismes,factor
    python scripts/exporter_excel.py --reprendre         # réutilise la dernière collecte
                                                         # (sortie/resultats-*.json), sans re-scraper

Dépendance : openpyxl (pip install openpyxl, inclus dans l'extra [dev]).
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from veille import extracteur  # noqa: E402
from veille.config import DOSSIER_SORTIE, FICHIER_SOURCES, FUSEAU_HORAIRE, charger_config  # noqa: E402
from veille.dedoublonnage import fusionner, trier  # noqa: E402
from veille.models import COLONNES, ProgrammeExtrait  # noqa: E402

logger = logging.getLogger("veille.excel")

ENTETES_JOURNAL = ["source", "nom", "programmes", "rejets", "avertissements", "erreur", "durée (s)"]


def _collecter(ids_demandes: str | None, delai_s: float):
    """Collecte fraîche sur les sites réels (le même chemin que veille.main)."""
    sources = [s for s in extracteur.charger_sources(FICHIER_SOURCES) if s.actif]
    if ids_demandes:
        demandes = {s.strip() for s in ids_demandes.split(",") if s.strip()}
        inconnues = demandes - {s.id for s in sources}
        if inconnues:
            raise SystemExit(f"Sources inconnues ou inactives : {', '.join(sorted(inconnues))}")
        sources = [s for s in sources if s.id in demandes]

    resultats = []
    for source in sources:
        logger.info("Scraping : %s (%d page(s))", source.nom, len(source.pages))
        resultats.append(extracteur.collecter_source(source, date.today(), delai_s))
    return resultats


def _reprendre_derniere_collecte():
    """Recharge la dernière sauvegarde JSON au lieu de re-scraper les sites."""
    fichiers = sorted(DOSSIER_SORTIE.glob("resultats-*.json"))
    if not fichiers:
        raise SystemExit(
            "Aucun fichier sortie/resultats-*.json — lancez d'abord une collecte "
            "(sans --reprendre, ou via python -m veille.main --console)."
        )
    contenu = json.loads(fichiers[-1].read_text(encoding="utf-8"))
    logger.info("Reprise de la collecte du %s (%s)", contenu["date"][:10], fichiers[-1].name)

    from types import SimpleNamespace

    class _Resultat:  # même surface que ResultatSource pour ce script
        def __init__(self, brut):
            self.source = SimpleNamespace(id=brut["id"], nom=brut["nom"])
            self.erreur = brut["erreur"]
            self.avertissements = brut["avertissements"]
            self.rejets = brut["rejets"]
            self.duree_s = brut["duree_s"]
            # Déjà validés à la collecte : on reconstruit sans re-valider.
            self.programmes = [ProgrammeExtrait.model_construct(**p) for p in brut["programmes"]]

    return [_Resultat(s) for s in contenu["sources"]]


def _ecrire_excel(chemin: Path, lignes, resultats) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    gras_blanc = Font(bold=True, color="FFFFFF")
    fond_entete = PatternFill("solid", fgColor="4472C4")
    fond_nouveau = PatternFill("solid", fgColor="C6EFCE")   # vert doux
    fond_urgent = PatternFill("solid", fgColor="FFD9B3")    # orange : échéance ≤ 14 jours
    aujourdhui = date.today()

    wb = Workbook()

    # ── Onglet Subventions ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Subventions"
    ws.append(COLONNES)
    for cellule in ws[1]:
        cellule.font = gras_blanc
        cellule.fill = fond_entete
    ws.freeze_panes = "A2"

    col_date = COLONNES.index("date_limite") + 1
    col_url = COLONNES.index("url") + 1
    col_statut = COLONNES.index("statut") + 1
    for ligne in lignes:
        ws.append(ligne.en_liste())
        rang = ws.max_row
        cellule_url = ws.cell(row=rang, column=col_url)
        if cellule_url.value:
            cellule_url.hyperlink = cellule_url.value
            cellule_url.font = Font(color="0563C1", underline="single")
        if ws.cell(row=rang, column=col_statut).value == "Nouveau":
            ws.cell(row=rang, column=col_statut).fill = fond_nouveau
        try:
            echeance = date.fromisoformat(str(ws.cell(row=rang, column=col_date).value))
            if 0 <= (echeance - aujourdhui).days <= 14:
                ws.cell(row=rang, column=col_date).fill = fond_urgent
        except ValueError:
            pass  # vide ou « continu »

    largeurs = {"nom_programme": 55, "organisme": 30, "montant": 22, "date_limite": 12,
                "url": 60, "notes_agent": 50, "admissibilite_obnl": 16}
    for i, nom in enumerate(COLONNES, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largeurs.get(nom, 14)
    ws.auto_filter.ref = ws.dimensions

    # ── Onglet Journal (une ligne par source) ─────────────────────────────────
    wj = wb.create_sheet("Journal")
    wj.append(ENTETES_JOURNAL)
    for cellule in wj[1]:
        cellule.font = gras_blanc
        cellule.fill = fond_entete
    wj.freeze_panes = "A2"
    for r in resultats:
        wj.append([
            r.source.id, r.source.nom, len(r.programmes), len(r.rejets),
            "; ".join(str(a) for a in r.avertissements), r.erreur or "", round(r.duree_s, 1),
        ])
    for i, largeur in enumerate([30, 40, 12, 8, 60, 40, 10], start=1):
        wj.column_dimensions[get_column_letter(i)].width = largeur
    for rang in wj.iter_rows(min_row=2):
        rang[4].alignment = Alignment(wrap_text=True)

    wb.save(chemin)


def principal() -> int:
    analyseur = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    analyseur.add_argument("--sources", metavar="IDS",
                           help="ids de sources, séparés par des virgules (voir sources.json)")
    analyseur.add_argument("--reprendre", action="store_true",
                           help="réutiliser la dernière collecte sauvegardée au lieu de re-scraper")
    args = analyseur.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s — %(message)s")

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit("openpyxl manquant — installez-le : pip install openpyxl")

    if args.reprendre:
        resultats = _reprendre_derniere_collecte()
    else:
        config = charger_config()
        resultats = _collecter(args.sources, config.delai_entre_requetes_s)

    programmes = [p for r in resultats for p in r.programmes]
    fusion = fusionner([], programmes, date.today())
    trier(fusion.lignes)

    DOSSIER_SORTIE.mkdir(exist_ok=True)
    maintenant = datetime.now(FUSEAU_HORAIRE)
    chemin = DOSSIER_SORTIE / f"veille-subventions-{maintenant.date().isoformat()}.xlsx"
    _ecrire_excel(chemin, fusion.lignes, resultats)

    en_erreur = [r.source.id for r in resultats if r.erreur]
    print()
    print(f"✔ {chemin}")
    print(f"  {len(fusion.lignes)} subventions uniques ({len(programmes)} extractions brutes)"
          f" — {len(resultats)} sources, {len(en_erreur)} en erreur"
          + (f" ({', '.join(en_erreur)})" if en_erreur else ""))
    print("  Ouvrez le fichier dans Excel pour vérifier : onglet Subventions (trié comme le"
          " futur Google Sheet) + onglet Journal (détail par source).")
    return 0


if __name__ == "__main__":
    sys.exit(principal())
